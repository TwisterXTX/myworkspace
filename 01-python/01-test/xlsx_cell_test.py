# # -*- coding: utf-8 -*-
# from openpyxl import Workbook
#
# wb = Workbook()
#
# ws1 = wb.create_sheet("Mysheet")
# ws1["A1"] = 123.11
# ws1["B2"] = "测试用例"
#
# # 行号和列号必须从1开始，获取指定单元格内容
# d = ws1.cell(row=1, column=1, value=10)
#
# # 指定行列坐标组成的字符串，value属性就是单元格值
# print(ws1["A1"].value)
# print(ws1["B2"].value)
# print(d.value)
# print(ws1.cell(row=4, column=2).value)
#
# # save file
# wb.save("xlsx_cell_test.xlsx")


# # -*- coding: utf-8 -*-
# from openpyxl import Workbook, load_workbook
# wb = Workbook()
# ws1 = wb.create_sheet("Mysheet")           #创建一个sheet
# ws1["A1"] = 1
# ws1["A2"] = 2
# ws1["A3"] = 3
# ##ws1["A10"]=3
# ws1["B1"] = 4
# ws1["B2"] = 5
# ws1["B3"] = 6
# ws1["C1"] = 7
# ws1["C2"] = 8
# ws1["C3"] = 9
# #操作单列
# print(ws1["A"])
# for cell in ws1["A"]:
#     print(cell.value)
#
# #最大有效行和列
# print(ws1.max_row, ws1.max_column)
# #最小有效行和列
# print(ws1.min_row, ws1.min_column)
# print(ws1.columns)
# print(ws1.rows)
# #操作单列
# print(ws1["A"])
# for cell in ws1["A"]:
#     print(cell.value)
#
# #操作多列,获取每一个值
# print(ws1["A:C"])
# for column in ws1["A:C"]:
#     for cell in column:
#         print(cell.value)
#
# #操作多行
# row_range = ws1[1:3]
# print(row_range)
# for row in row_range:
#     for cell in row:
#         print(cell.value)
#
# #指定一个操作的区域
# print("*"*50)
# for row in ws1.iter_rows(min_row=1, min_col=1, max_col=3, max_row=3):
#     for cell in row:
#         print(cell.value)
#
# #获取所有行
# print(ws1.rows)
# for row in ws1.rows:
#     print(row)
#     print("*"*50)
#
# #获取所有列
# print(ws1.columns)
# for col in ws1.columns:
#     print(col)
#
# # 保存文件
# wb.save("xlsx_cell_test1.xlsx")

# -*- coding: utf-8 -*-
# 合并/取消合并单元格
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

ws.merge_cells('A2:D2')#合并单元格
ws.unmerge_cells('A2:D2')#取消合并
ws['A2'] = "hello world"
# 或者
# ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)
# ws.unmerge_cells(start_row=2,start_column=1,end_row=2,end_column=4)

wb.save('xlsx_cell_test2.xlsx')

