#!/usr/bin/python
# -*- coding: UTF-8 -*-
import cv2
import openpyxl
import time
import datetime
import locale

wb = openpyxl.Workbook()

ws = wb.active

ws['A1'] = 42
ws['B1'] = "习天翔"+"is handsome"
ws['A2'] = datetime.datetime.now()

locale.setlocale(locale.LC_CTYPE, 'chinese')
ws['A3'] = time.strftime("%Y年%m月%d日 %H时%M分%S秒", time.localtime())

ws.append([1, 2, 3])

wb.save('a.xlsx')

# print(time.localtime())


# img = cv2.imread('bird.jpg')
#
# cv2.imshow('imshow', img)
#
# cv2.waitKey(0)
#
# cv2.destroyAllWindows()
