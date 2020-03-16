# -*- coding: utf-8 -*-
'''
插入图片
'''

# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image
#
# wb = load_workbook('a.xlsx')
# ws1 = wb.active
# img = Image('./picture/bird.jpg')
# ws1.add_image(img, 'A2')
#
# wb.save("a.xlsx")

'''
插入图片
'''

import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as pyxlImage
import PIL.Image as Image
import cv2

IMAGE_PATH = './picture/'
IMAGES_FORMAT = ['.jpg', '.JPG']
IMAGE_SIZE = 256
IMAGE_ROW = 2
IMAGE_COL = 5
IMAGE_SAVE_PATH = 'final.jpg'

# get all active pictures in the folder
image_name = [name for name in os.listdir(IMAGE_PATH) for item in IMAGES_FORMAT if
              os.path.splitext(name)[1] == item and 'img' in os.path.splitext(name)[0]]

if len(image_name) != IMAGE_ROW * IMAGE_COL:
    raise ValueError("image para and num are not match")


# def image compose function
def image_compose():
    to_image = Image.new('RGB', (IMAGE_COL * IMAGE_SIZE, IMAGE_ROW * IMAGE_SIZE))  # create a new image
    # go through all picture and paste at assigned position
    for y in range(1, IMAGE_ROW + 1):
        for x in range(1, IMAGE_COL + 1):
            from_image = Image.open(IMAGE_PATH + image_name[IMAGE_COL * (y - 1) + x - 1]).resize(
                (IMAGE_SIZE, IMAGE_SIZE), Image.ANTIALIAS)
            to_image.paste(from_image, ((x - 1) * IMAGE_SIZE, (y - 1) * IMAGE_SIZE))
    return to_image.save(IMAGE_SAVE_PATH)


image_compose()
wb = load_workbook('a.xlsx')
ws1 = wb.active
img = pyxlImage('final.jpg')
ws1.add_image(img, 'A2')

wb.save("a.xlsx")

img_cv = cv2.imread('final.jpg')

cv2.imshow('imshow', img_cv)

cv2.waitKey(0)

cv2.destroyAllWindows()
