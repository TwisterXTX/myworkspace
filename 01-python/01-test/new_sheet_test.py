# -*- coding: utf-8 -*-
from openpyxl import Workbook

wb = Workbook()

ws = wb.create_sheet("kongsh")
ws1 = wb.create_sheet("Mysheet")

ws1.title = "New Title"
ws2 = wb.create_sheet("Mysheet", 0)     #设定sheet的插入位置
ws2.title = u"测试用例"
# 设置sheet名的背景色
ws1.sheet_properties.tabColor = "1072BA"

print(wb["测试用例"])
print(wb["New Title"])

print(wb.sheetnames)
for sheet_name in wb.sheetnames:
    print(sheet_name)
    print(wb[sheet_name])

# 复制一个sheet
wb["New Title"]["A1"] = "gloryroad"
source = wb["New Title"]
target = wb.copy_worksheet(source)

target.title = "New copy Title"

wb.save("new_sheet_test.xlsx")
